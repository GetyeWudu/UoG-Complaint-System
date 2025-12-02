"""
Reporting and export functionality for complaints
Supports PDF and Excel export
"""
from .models import Complaint, ComplaintEvent, Category
from accounts.models import CustomUser, Department, College, Campus
from django.db.models import Q, Count, Avg, F, Sum
from django.utils import timezone
from datetime import timedelta
import csv
import json
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch


def generate_complaints_report(complaints_queryset, format='excel', filters=None):
    """
    Generate a report of complaints in Excel or PDF format.
    
    Args:
        complaints_queryset: QuerySet of complaints
        format: 'excel' or 'pdf'
        filters: Dict of applied filters
    
    Returns:
        BytesIO object with report data
    """
    if format == 'excel':
        return generate_excel_report(complaints_queryset, filters)
    elif format == 'pdf':
        return generate_pdf_report(complaints_queryset, filters)
    else:
        raise ValueError(f"Unsupported format: {format}")


def generate_excel_report(complaints_queryset, filters=None):
    """Generate Excel report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Complaints Report"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = [
        'Tracking ID', 'Title', 'Status', 'Priority', 'Category', 
        'Submitter', 'Assigned To', 'Location', 'Created At', 
        'Resolved At', 'Resolution Time (hours)', 'Rating'
    ]
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for complaint in complaints_queryset:
        resolution_time = complaint.time_to_resolution()
        row = [
            complaint.tracking_id,
            complaint.title[:50],  # Truncate long titles
            complaint.get_status_display(),
            complaint.get_priority_display(),
            complaint.category.name if complaint.category else 'N/A',
            complaint.submitter.get_full_name() if complaint.submitter else 'Anonymous',
            complaint.assigned_to.get_full_name() if complaint.assigned_to else 'Unassigned',
            complaint.location,
            complaint.created_at.strftime('%Y-%m-%d %H:%M'),
            complaint.resolved_at.strftime('%Y-%m-%d %H:%M') if complaint.resolved_at else 'N/A',
            f"{resolution_time:.1f}" if resolution_time else 'N/A',
            complaint.feedback_rating if complaint.feedback_rating else 'N/A'
        ]
        ws.append(row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_pdf_report(complaints_queryset, filters=None):
    """Generate PDF report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Center
    )
    story.append(Paragraph("Complaints Report", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Filters info
    if filters:
        filter_text = "Filters: " + ", ".join([f"{k}: {v}" for k, v in filters.items()])
        story.append(Paragraph(filter_text, styles['Normal']))
        story.append(Spacer(1, 0.1*inch))
    
    # Summary stats
    total = complaints_queryset.count()
    resolved = complaints_queryset.filter(status__in=['resolved', 'closed']).count()
    story.append(Paragraph(f"Total Complaints: {total} | Resolved: {resolved}", styles['Normal']))
    story.append(Spacer(1, 0.2*inch))
    
    # Table data
    data = [['Tracking ID', 'Title', 'Status', 'Priority', 'Created']]
    
    for complaint in complaints_queryset[:100]:  # Limit to 100 for PDF
        data.append([
            complaint.tracking_id,
            complaint.title[:40],
            complaint.get_status_display(),
            complaint.get_priority_display(),
            complaint.created_at.strftime('%Y-%m-%d')
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.2*inch))
    story.append(Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def get_dashboard_statistics(user, role, date_range=None):
    """
    Get comprehensive statistics for dashboard based on user role.
    Returns dict with various metrics.
    """
    if date_range is None:
        date_range = (timezone.now() - timedelta(days=30), timezone.now())
    
    start_date, end_date = date_range
    
    # Base queryset based on role
    if role == 'student':
        complaints = Complaint.objects.filter(submitter=user)
    elif role == 'dept_head':
        if user.department:
            complaints = Complaint.objects.filter(department=user.department)
        else:
            complaints = Complaint.objects.none()
    elif role == 'dean':
        if user.department and user.department.college:
            departments = Department.objects.filter(college=user.department.college)
            complaints = Complaint.objects.filter(department__in=departments)
        else:
            complaints = Complaint.objects.none()
    elif role == 'campus_director':
        if user.campus:
            complaints = Complaint.objects.filter(campus=user.campus)
        else:
            complaints = Complaint.objects.none()
    elif role in ['admin', 'super_admin']:
        complaints = Complaint.objects.all()
    else:
        complaints = Complaint.objects.filter(assigned_to=user)
    
    # Filter by date range
    complaints = complaints.filter(created_at__range=date_range)
    
    # Calculate statistics
    stats = {
        'total': complaints.count(),
        'new': complaints.filter(status='new').count(),
        'assigned': complaints.filter(status='assigned').count(),
        'in_progress': complaints.filter(status='in_progress').count(),
        'resolved': complaints.filter(status='resolved').count(),
        'closed': complaints.filter(status='closed').count(),
        'rejected': complaints.filter(status='rejected').count(),
        'by_priority': {},
        'by_category': {},
        'sla_breaches': complaints.filter(
            Q(sla_response_breached=True) | Q(sla_resolution_breached=True)
        ).count(),
        'average_resolution_time': None,
        'average_rating': None,
    }
    
    # By priority
    for priority in ['critical', 'high', 'medium', 'low']:
        stats['by_priority'][priority] = complaints.filter(priority=priority).count()
    
    # By category
    category_stats = complaints.values('category__name').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    stats['by_category'] = {item['category__name'] or 'Uncategorized': item['count'] 
                           for item in category_stats}
    
    # Average resolution time
    resolved_complaints = complaints.filter(status__in=['resolved', 'closed'])
    if resolved_complaints.exists():
        resolution_times = [
            c.time_to_resolution() for c in resolved_complaints 
            if c.time_to_resolution() is not None
        ]
        if resolution_times:
            stats['average_resolution_time'] = sum(resolution_times) / len(resolution_times)
    
    # Average rating
    rated = complaints.exclude(feedback_rating__isnull=True)
    if rated.exists():
        stats['average_rating'] = rated.aggregate(Avg('feedback_rating'))['feedback_rating__avg']
    
    return stats

